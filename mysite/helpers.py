import mws, os
from datetime import date
from mws import Orders, Products, Inventory
import json
import requests
from bs4 import BeautifulSoup

from mysite.functions.functions import parseAmazon, parseInfo, getInfos, getFees

import time
from openpyxl import load_workbook


import pymongo




myclient = pymongo.MongoClient("mongodb://localhost:27017/")

mydb = myclient["amazon"]
mycol = mydb["characters"]


access_key="AKIAJA2FY3FIHVRFF2NQ"
secret_key="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"
account_id="A2IE91UMYBQVEE"
region='US'


def ListOrders():

    api = Orders(
        access_key=access_key,
        secret_key=secret_key,
        account_id=account_id,
        region=region,    
    )

    api._use_feature_mwsresponse = True

    response = api.list_orders(marketplaceids="ATVPDKIKX0DER",created_after="2014-07-06T19:00:00Z")

    orders = []
    for order in response.parsed.Orders.Order:
        orders.append(order.AmazonOrderId)
        break
       
    result = api.get_order(amazon_order_ids=orders[0])

    print(result)
     


def ListProducts(query):
    api = Products(
        access_key=access_key,
        secret_key=secret_key,
        account_id=account_id,
        region=region,  
    )

    api._use_feature_mwsresponse = True

    response = api.list_matching_products(marketplaceid="ATVPDKIKX0DER",query=query)
    print(response.parsed)

import os.path


def GetProductInfo(asin,parsed,show):
   
    api = Products(
        access_key=access_key,
        secret_key=secret_key,
        account_id=account_id,
        region=region,  
    )

    print("######")
    if os.path.exists('./1.txt'):
        print("ASDASDASD")
    else:
        f = open("./1.txt", "a")
        f.write(parsed.Product)
        f.close()

        print(parsed.Product)
  
    height = parsed.Product.AttributeSets.ItemAttributes.ItemDimensions.Height.value
    weight = parsed.Product.AttributeSets.ItemAttributes.ItemDimensions.Weight.value
    length = parsed.Product.AttributeSets.ItemAttributes.ItemDimensions.Length.value
    width = parsed.Product.AttributeSets.ItemAttributes.ItemDimensions.Width.value
    
    fees = getFees(asin,height,weight,length,width)

    api._use_feature_mwsresponse = True


    fulfillment = api.get_my_price_for_asin(marketplaceid="ATVPDKIKX0DER",asins=asin)
    result = api.get_competitive_pricing_for_asin(marketplaceid="ATVPDKIKX0DER",asins=asin)


    parseInfo(asin)


    contents = ""


    with open('result.txt', 'r', encoding="utf8") as f:

        contents = f.read()


    infos = getInfos(contents)
    
    seller_rank = ""
    for rank in infos[4]:
        seller_rank += rank+" "
    

    parseAmazon(infos[0],infos[1],infos[2])
    
    contents = ""
    with open('asd.htm', 'r', encoding="utf8") as f:

        contents = f.read()

    amount = ''
    start = False
    

    for i in range(0,len(contents)):
        if start==False and contents[i]=='>' and contents[i+1]==',' and contents[i+3]=='t' and contents[i+4]=='h':
            start = True
            i+=12

        if start==True:
            amount += contents[i]
        
        if start==True and contents[i+1]=='<':
            break
    
    try:
        amount = amount.slit(",")[0]
    except:
        print("")
    
    #amount = amount.split(",")[0]

    if show==True:
        data = {"Inventory":amount,"Price":infos[3],"SellerRank":seller_rank,'Image':infos[5],'SellersName':infos[6],'ASIN':asin,'Parsed':parsed}
    else:
        data = {"Inventory":amount,"Price":infos[3],"SellerRank":seller_rank,'Image':infos[5],'SellersName':infos[6],'Fees':json.loads(fees),'CompetitivePricing':result.parsed}
    return data



def GetInventory(sellerSKU):

    api = Inventory(
        access_key=access_key,
        secret_key=secret_key,
        account_id=account_id,
        region=region,  
    )

    api._use_feature_mwsresponse = True

    response = api.list_inventory_supply(skus=sellerSKU)
    print(response.parsed)




def GetProduct(asin):

    api = Products(
        access_key=access_key,
        secret_key=secret_key,
        account_id=account_id,
        region=region,  
    )

    #https://www.onlinejobs.ph/jobseekers/job/470624
    #https://www.onlinejobs.ph/jobseekers/job/470512

    api._use_feature_mwsresponse = True

    #result = api.get_competitive_pricing_for_asin(marketplaceid="ATVPDKIKX0DER",asins=asin)
    #return result.parsed
    tmp = api.list_matching_products(marketplaceid="ATVPDKIKX0DER",query="017754398666")
   
    #print(tmp.parsed)
    
    response1 = api.get_lowest_offer_listings_for_asin(marketplaceid="ATVPDKIKX0DER",asins=asin)
    response = api.get_matching_product(marketplaceid="ATVPDKIKX0DER",asins=asin)

    print("@#@#")


  
    #print(response.parsed)

    test= {}
    similar = {}
    try:
        test = response.parsed.Product.Relationships.VariationParent.Identifiers.MarketplaceASIN.ASIN
        similar = GetProductInfo(test,response.parsed,False)
    except:
        print("DDDDD")

    return {'Original':GetProductInfo(asin,response.parsed,True),'Similar':similar,'Data':tmp.parsed}
       


def RetrieveProducts():
    myquery = { "ip": "1111" }

    mydoc = mycol.find_one(myquery)
    return mydoc["results"]


def GetProductFromJS(upc):
    

    api = Products(
        access_key=access_key,
        secret_key=secret_key,
        account_id=account_id,
        region=region,  
    )


    api._use_feature_mwsresponse = True

    #result = api.get_competitive_pricing_for_asin(marketplaceid="ATVPDKIKX0DER",asins=asin)
    #return result.parsed
    

    Original = []
    Similar = []
    response = ""
    tmp  = ""
    data = {}
    fba = 0
    merch = 0

    try:

        tmp = api.list_matching_products(marketplaceid="ATVPDKIKX0DER",query=upc)
      
        the_asin = tmp.parsed.Products.Product.Identifiers.MarketplaceASIN.ASIN
    
        response = api.get_matching_product(marketplaceid="ATVPDKIKX0DER",asins=the_asin)

        response1 = api.get_lowest_offer_listings_for_asin(marketplaceid="ATVPDKIKX0DER",asins=the_asin)

     

        for i in range(0,len(response1.parsed.Product.LowestOfferListings.LowestOfferListing)):
            if "Amazon" in response1.parsed.Product.LowestOfferListings.LowestOfferListing[i].Qualifiers.FulfillmentChannel:
                fba += 1
            else:
                merch +=1

        Original = GetProductInfo(the_asin,response.parsed,True)



    except Exception as ee:
        print(ee)


    try:

       
        test = response.parsed.Product.Relationships.VariationParent.Identifiers.MarketplaceASIN.ASIN
       
        Similar = GetProductInfo(test,response.parsed,False)
        
    
    except Exception as ee:
        print(ee)



    return {'Original':Original,'Similar':Similar,'Data':{'FBA':fba,'Merchant':merch}}
       


def GetProductData(id,ip):
    dblist = myclient.list_database_names()
    if "amazon" in dblist:
        print("The database exists.")
    myquery = { "ip": ip }

    mydoc = mycol.find_one(myquery)

    return mydoc["results"][id]




def GetProductResults(ip):
    dblist = myclient.list_database_names()
    if "amazon" in dblist:
        print("The database exists.")
    myquery = { "ip": ip }

    mydoc = mycol.find_one(myquery)
    #print(mydoc["results"])
    return mydoc["results"]




def handle_uploaded_file(f,ip):

    dblist = myclient.list_database_names()
    if "amazon" in dblist:
        print("The database exists.")


    fn = 'mysite/static/upload/'+f.name
    with open(fn, 'wb+') as destination:  
        for chunk in f.chunks():  
            destination.write(chunk)  
    
    #https://mdbootstrap.com/plugins/jquery/video/
    wb = load_workbook(filename = fn)
    ws = wb.active

    allUPCS = ""
    data = {}

    for i in range(2,len(ws['C'])):
        try:
            
            data = GetProductFromJS(ws['C'+str(i)].value)

          
            
        except Exception as ee:
            print(ee)
        

        myquery = { "ip": ip }

        mydoc = mycol.find_one(myquery)
        print("HERE>>>>")
        #print(mydoc)

        if mydoc == None:
            print("TRYING TO ADD")
            mycol.insert_one({"ip":ip,"results":[]})

        else:
            print("FROM MONGO")
            mycol.update(myquery, {'$push': {'results': data}})
            #mydoc["results"].insert_one(data)
            #print(mydoc)

   
 

    return allUPCS

#GetProduct("B07RF1XD36")
